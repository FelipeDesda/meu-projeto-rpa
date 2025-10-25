from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb= load_workbook('relatorio.xlsx')
ws = wb.active

valor = ws['B2'].value
ws['C5'] = 'Atualizado'

ws['A1'].font = Font(bold=True, size=14)
ws['A1'].fill = PatternFill(start_color='FFFF00', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center')

nova_planilha = wb.create_sheet("Resumo")
nova_planilha['A1'] = 'Relatório de Vendas'

wb.save('relatorio_formatado.xlsx')